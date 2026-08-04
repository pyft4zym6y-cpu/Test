#!/usr/bin/env python3
"""Заполняет колонку «ОТВЕТ КЛИЕНТА» в ЕКП-книге (XLSX) ответами из портала.

Использование:
    python fill_ekp.py ekp-answers-<client>.csv "Consulting Discovery Framework v2.0.xlsx" [выход.xlsx]

CSV берётся из портала: /admin/c/<клиент> → «Экспорт для ЕКП (CSV)».
Скрипт ищет ID вопроса (первая колонка листа, формат XX-123) на всех листах
книги и пишет ответ в колонку «ОТВЕТ КЛИЕНТА» (ищется по заголовку в первых
пяти строках листа; если не нашлась — колонка 8). Исходный файл не трогается,
результат сохраняется в новую книгу.

Зависимость: pip install openpyxl
"""
import csv
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    sys.exit("Нужен openpyxl: pip install openpyxl")

ID_RE = re.compile(r"^[A-ZА-ЯІЇЄ]{2,3}-\d{1,4}$")
ANSWER_HEADER_RE = re.compile(r"ОТВЕТ|ВІДПОВІДЬ", re.IGNORECASE)
DEFAULT_ANSWER_COL = 8
FACTS_SUFFIX = True  # дописывать факты/ссылки клиента в скобках после ответа


def read_answers(csv_path: Path) -> dict[str, tuple[str, str]]:
    answers: dict[str, tuple[str, str]] = {}
    with csv_path.open(encoding="utf-8-sig", newline="") as f:
        sample = f.read(4096)
        f.seek(0)
        delim = ";" if sample.count(";") >= sample.count(",") else ","
        for row in csv.DictReader(f, delimiter=delim):
            qid = (row.get("ID") or "").strip()
            ans = (row.get("ОТВЕТ КЛИЕНТА") or "").strip()
            facts = (row.get("Факты") or "").strip()
            if qid and ans:
                answers[qid] = (ans, facts)
    return answers


def find_answer_col(ws) -> int:
    for row in ws.iter_rows(min_row=1, max_row=5):
        for cell in row:
            if isinstance(cell.value, str) and ANSWER_HEADER_RE.search(cell.value):
                return cell.column
    return DEFAULT_ANSWER_COL


def main() -> None:
    if len(sys.argv) < 3:
        sys.exit(__doc__)
    csv_path, xlsx_path = Path(sys.argv[1]), Path(sys.argv[2])
    out_path = Path(sys.argv[3]) if len(sys.argv) > 3 else xlsx_path.with_stem(xlsx_path.stem + " — заполнено")

    answers = read_answers(csv_path)
    if not answers:
        sys.exit("В CSV нет ответов (колонки ID / ОТВЕТ КЛИЕНТА).")
    print(f"Ответов в CSV: {len(answers)}")

    wb = load_workbook(xlsx_path)
    filled, seen = 0, set()
    for ws in wb.worksheets:
        col = find_answer_col(ws)
        for row in ws.iter_rows(min_row=1):
            qid = row[0].value
            if not isinstance(qid, str):
                continue
            qid = qid.strip()
            if not ID_RE.match(qid) or qid not in answers:
                continue
            ans, facts = answers[qid]
            value = f"{ans} ({facts})" if FACTS_SUFFIX and facts else ans
            ws.cell(row=row[0].row, column=col, value=value)
            filled += 1
            seen.add(qid)

    missing = sorted(set(answers) - seen)
    wb.save(out_path)
    print(f"Заполнено ячеек: {filled} → {out_path}")
    if missing:
        print(f"Не нашлись в книге ({len(missing)}): {', '.join(missing[:20])}{'…' if len(missing) > 20 else ''}")


if __name__ == "__main__":
    main()
